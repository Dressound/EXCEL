import xlrd
import xlsxwriter
from tkinter import filedialog, simpledialog, simpledialog, filedialog, tk
import re
from datetime import datetime as dt
import pandas as pd
import openpyxl
import os

# Definir la fecha en la que deseas bloquear el programa
fecha_bloqueo = dt(2025, 12, 1)

# Verificar si la fecha actual es posterior a la fecha de bloqueo
if dt.now() > fecha_bloqueo:
    print("El programa está bloqueado. Por favor, contacta al administrador 0993842259")
else:
    contraseña = simpledialog.askstring("Contraseña", "Introduce la contraseña:")

    if contraseña == "talentohumano":
        print("¡Contraseña correcta! El programa está desbloqueado y en funcionamiento.")

        def extraer_texto_desde_excel(archivo_excel, nombre_hoja):
            try:
                libro_excel = xlrd.open_workbook(archivo_excel)
                hoja = libro_excel.sheet_by_name(nombre_hoja)
                filas_texto = []
                for fila_idx in range(hoja.nrows):
                    fila_texto = []
                    for col_idx in range(hoja.ncols):
                        celda_valor = hoja.cell_value(fila_idx, col_idx)
                        fila_texto.append(celda_valor)
                    filas_texto.append(fila_texto)
                return filas_texto
            except Exception as e:
                print("Error al extraer texto desde el archivo Excel:", e)
                return None

        def procesar_fecha_hora_am_pm(fecha_hora_am_pm):
            try:
                patron = r"(\d{1,2}/\d{1,2}/\d{2,4}) (\d{1,2}:\d{2}) ([APap]\.?[Mm]\.?)"
                coincidencias = re.findall(patron, fecha_hora_am_pm)
                if coincidencias:
                    fecha, hora, am_pm = coincidencias[0]
                    return fecha, hora, am_pm.upper()
                else:
                    raise ValueError("No se pudo encontrar el formato de fecha, hora y AM/PM en la cadena.")
            except ValueError as ve:
                print(ve)
                return None, None, None

        def convertir_a_fecha_larga_espanol(fecha_str):
            if fecha_str:
                try:
                    dia, mes, año = fecha_str.split('/')
                    if len(año) == 2:
                        if int(año) >= 0 and int(año) <= 30:
                            año = '20' + año
                        else:
                            año = '19' + año
                    fecha_str = f"{dia}/{mes}/{año}"
                    fecha = dt.strptime(fecha_str, '%d/%m/%Y')
                    
                    dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
                    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
                    
                    dia_semana = dias_semana[fecha.weekday()]
                    mes = meses[fecha.month - 1]
                    fecha_larga = f"{dia_semana}, {fecha.day:02d} de {mes} de {fecha.year}"
                    return fecha_larga
                except ValueError as ve:
                    print("Error al convertir la fecha:", ve)
            return None

        def procesar_texto(texto):
            texto_procesado = []
            procesando_empleado = False
            for fila in texto:
                if "EMPLEADO" in fila:
                    procesando_empleado = True
                    texto_procesado.append(fila)
                elif procesando_empleado:
                    if len(fila) >= 2:
                        fecha_hora_am_pm = fila[7]
                        fecha, hora, am_pm = procesar_fecha_hora_am_pm(fecha_hora_am_pm)
                        if fecha is not None and hora is not None and am_pm is not None:
                            fecha_larga = convertir_a_fecha_larga_espanol(fecha)
                            hora_am_pm = f"{hora} {am_pm}"
                            fila.insert(8, fecha_larga)
                            fila.insert(9, hora_am_pm)
                            texto_procesado.append(fila)
                    if "EMPLEADO" in fila:
                        procesando_empleado = False
                        print("Empleado procesado.")
                else:
                    texto_procesado.append(fila)
            return texto_procesado

        def seleccionar_archivo():
            archivo_excel = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xls"), ("Todos los archivos", "*.*")])
            if archivo_excel:
                texto_extraido = extraer_texto_desde_excel(archivo_excel, "reporteTimbresDetalleGrupo")
                if texto_extraido:
                    texto_procesado = procesar_texto(texto_extraido)
                    guardar_archivo(texto_procesado)

        def guardar_archivo(texto):
            archivo_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")])
            if archivo_guardado:
                try:
                    libro_excel = xlsxwriter.Workbook(archivo_guardado, {'strings_to_urls': False, 'constant_memory': True, 'default_date_format': 'dd/mm/yy', 'strings_to_numbers': True, 'strings_to_formulas': True, 'encoding': 'utf-8'})
                    hoja = libro_excel.add_worksheet("TextoExtraido")
                    for fila_idx, fila in enumerate(texto):
                        for col_idx, valor_celda in enumerate(fila):
                            hoja.write(fila_idx, col_idx, valor_celda)
                    libro_excel.close()
                    print("Archivo guardado correctamente.")
                except Exception as e:
                    print("Error al guardar el archivo Excel:", e)

        ventana = tk.Tk()
        ventana.title("Extractor de texto de Excel")
        boton_seleccionar = tk.Button(ventana, text="Seleccionar archivo Excel", command=seleccionar_archivo)
        boton_seleccionar.pack(pady=10)
        ventana.mainloop()

    else:
        print("Contraseña incorrecta. El programa se cerrará.")

# Función para seleccionar el archivo
def select_file(title):
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal
    file_path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx *.xls")])
    return file_path

# Función para cargar el archivo Excel (soporta .xls y .xlsx)
def load_excel(file_path, sheet_name=None):
    try:
        if file_path.endswith('.xls'):
            print(f"Cargando archivo {file_path} con xlrd")
            return pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='xlrd')
        else:
            print(f"Cargando archivo {file_path} con pandas")
            return pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    except Exception as e:
        print(f"Error al cargar el archivo {file_path}: {e}")
        return None

# Función para hacer preguntas al usuario
def ask_user_input(prompt):
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal
    user_input = simpledialog.askstring("Input", prompt)
    return user_input

# Seleccionar el archivo base
base_file = select_file("Selecciona el archivo base")

# Asignar nombres de hoja por defecto
base_sheet_name = "TextoExtraido"

# Cargar los datos del archivo base
base_df = load_excel(base_file, sheet_name=base_sheet_name)

# Verificar si el archivo base se ha cargado correctamente
if base_df is None:
    print("Error: No se pudo cargar el archivo base.")
    exit()

# Preguntar al usuario el número de la columna (1-indexado)
column_empleado = int(ask_user_input("¿En qué número de columna está 'empleado'? (Ej. 2 para la columna B)")) - 1
column_nombre = int(ask_user_input("¿En qué número de columna está el nombre del empleado? (Ej. 8 para la columna H)")) - 1
column_fecha = int(ask_user_input("¿En qué número de columna están las fechas? (Ej. 9 para la columna I)")) - 1
column_hora = int(ask_user_input("¿En qué número de columna están las horas? (Ej. 10 para la columna J)")) - 1
column_entrada_salida = int(ask_user_input("¿En qué número de columna están las entradas y salidas? (Ej. 18 para la columna R)")) - 1

# Inicializar la lista de empleados con errores
errores = []

# Inicializar el nuevo archivo de Excel
nuevo_archivo = "transformado.xlsx"
wb_nuevo = openpyxl.Workbook()
hoja_nuevo = wb_nuevo.active

# Insertar una nueva fila al principio e ingresar los encabezados
hoja_nuevo.insert_rows(1)
hoja_nuevo.cell(row=1, column=1, value="Fecha")
hoja_nuevo.cell(row=1, column=2, value="Hora")
hoja_nuevo.cell(row=1, column=3, value="Entrada/Salida")

# Inicializar la fila de inicio para pegar los datos
current_row = 2  # Comenzar desde la segunda fila para los datos

# Iterar sobre cada fila en el archivo base
for index, row in base_df.iterrows():
    if str(row.iloc[column_empleado]).strip().lower() == 'empleado':  # Insensible a mayúsculas y minúsculas
        print(f"Encontrado 'empleado' en la fila {index}")
        # Obtener el nombre del empleado
        try:
            nombre_empleado = base_df.iat[index, column_nombre]
            print(f"Nombre del empleado encontrado: {nombre_empleado}")
        except Exception as e:
            print(f"Error al obtener el nombre del empleado en la fila {index}: {e}")
            continue

        fechas = []
        horas = []
        entradas_salidas = []

        # Obtener las fechas, horas y entradas/salidas
        i = index + 1
        while i < len(base_df) and pd.notna(base_df.iat[i, column_fecha]):
            try:
                fechas.append(base_df.iat[i, column_fecha])
                horas.append(base_df.iat[i, column_hora])
                entradas_salidas.append(base_df.iat[i, column_entrada_salida])
                i += 1
            except Exception as e:
                print(f"Error al obtener datos en la fila {i}: {e}")
                break

        # Pegar el nombre del empleado
        try:
            hoja_nuevo.cell(row=current_row, column=1, value=nombre_empleado)
            print(f"Nombre del empleado {nombre_empleado} pegado en la fila {current_row}")
        except Exception as e:
            print(f"Error al pegar el nombre del empleado en la fila {current_row}: {e}")
            continue

        # Pegar las fechas, horas y entradas/salidas
        try:
            for j in range(len(fechas)):
                hoja_nuevo.cell(row=current_row + 1 + j, column=1, value=fechas[j])
                hoja_nuevo.cell(row=current_row + 1 + j, column=2, value=horas[j])
                hoja_nuevo.cell(row=current_row + 1 + j, column=3, value=entradas_salidas[j])
            print(f"Datos pegados para el empleado {nombre_empleado} desde la fila {current_row + 1}")
            current_row += len(fechas) + 2  # Actualizar la fila actual con un espacio adicional
        except Exception as e:
            print(f"Error al pegar datos para el empleado {nombre_empleado}: {e}")
            continue

# Guardar el archivo con los datos pegados
try:
    wb_nuevo.save(nuevo_archivo)
    print(f"Proceso completado. Se ha creado un nuevo archivo con los datos pegados como '{nuevo_archivo}'.")
except Exception as e:
    print(f"Error al guardar el nuevo archivo: {e}")

# Guardar nombres de empleados con errores en un archivo .txt
if len(errores) > 0:
    try:
        with open("empleados_con_errores.txt", "w") as f:
            for empleado in errores:
                f.write(empleado + "\n")
        print("Nombres de empleados con errores guardados en 'empleados_con_errores.txt'.")
    except Exception as e:
        print(f"Error al guardar los nombres de empleados con errores: {e}")
# Ruta al archivo Excel
file_path = "transformado.xlsx"

# Verificar si el archivo existe
if not os.path.isfile(file_path):
    raise FileNotFoundError(f"No such file or directory: '{file_path}'")

# Leer el archivo Excel
df = pd.read_excel(file_path)

# Función para determinar si se trabajó en un turno de velada
def es_velada(entrada, salida):
    entrada_str = str(entrada)
    salida_str = str(salida)
    if entrada_str[-2:] == "PM" and salida_str[-2:] == "AM":
        return True
    return False

# Variables para rastrear las veladas de cada empleado y los nombres repetidos
veladas_empleado = {}
nombres_repetidos = set()

# Iterar sobre cada fila del DataFrame
for index, row in df.iterrows():
    # Obtener el nombre del empleado
    empleado = row['Fecha']
    
    # Verificar si es una fila de empleado o una fila vacía
    if pd.isnull(empleado):
        continue
    else:
        if empleado in veladas_empleado:
            nombres_repetidos.add(empleado)
        else:
            veladas_empleado[empleado] = 0
    
    # Obtener los valores de la fila
    hora = row['Hora']
    tipo = row['Entrada/Salida']
    
    # Marcar el tipo y el turno
    if tipo == 'Entrada':
        if index + 1 < len(df) and es_velada(hora, df.iloc[index + 1]['Hora']):
            df.at[index, 'Tipo'] = 'V'
            df.at[index, 'Turno'] = ''
            veladas_empleado[empleado] += 1
        else:
            # Verificar si es una entrada consecutiva en la misma fecha para el mismo empleado
            if (index + 1 < len(df) and 
                df.iloc[index + 1]['Fecha'] == empleado and 
                df.iloc[index + 1]['Entrada/Salida'] == 'Entrada'):
                df.at[index + 1, 'Tipo'] = 'ERROR'
                df.at[index + 1, 'Turno'] = 'ERROR'
            else:
                df.at[index, 'Tipo'] = 'T'
                df.at[index, 'Turno'] = 'T'

# Guardar los nombres de empleados repetidos en un archivo de texto
with open("nombres_repetidos.txt", "w") as file:
    for nombre in nombres_repetidos:
        file.write(nombre + "\n")

# Guardar el DataFrame modificado en un nuevo archivo Excel
df.to_excel("nuevo_transformado.xlsx", index=False)
